import openpyxl

from api.api_key import ApiKey


class ExcelRead():
    def __init__(self, path):
        self.path = path

    def get_sheet_name(self):
        """
        获取tab的左右sheets
        :param path:
        :return: list sheets
        """
        work_book = openpyxl.load_workbook(self.path)
        return work_book.sheetnames

    def get_sheet_by_sheet_name(self, sheet_name):
        """
        获取指定名字的sheet表
        :param sheet_name:
        :return:
        """
        work_book = openpyxl.load_workbook(self.path)
        sheet = work_book[sheet_name]
        return sheet

    def send_requests(self, sheet_name):
        sheet = self.get_sheet_by_sheet_name(sheet_name)
        ak = ApiKey()
        r = 0
        for value in sheet.values:
            r = r + 1
            print("第{}行的数据是：".format(r))
            # print("类型是：{}".format(type(value)))

            if type(value[0]) is int:
                if value[4]:
                    # print("存在请求头")
                    if value[5]:
                        data_dic = {
                            "url": value[1]+value[2],
                            "headers": eval(value[4]),
                            "json": eval(value[5])
                        }

                        response = getattr(ak, value[3])(**data_dic)
                        result = ak.get_text(response.text, value[7])

                        try:
                            if result == value[8]:
                                # sheet.cell(r, 10).value = "通过"
                                print("通过")
                            else:

                                sheet.cell(r, 10).value = "不通过"
                                print("不通过：result是{}，期望是：{}".format(result, value[8]))
                        except Exception as e:
                            print('error:', e)
                    else:
                        data_dic = {
                            "url": value[1] + value[2],
                            "headers": eval(value[4])
                        }
                        response = getattr(ak, value[3])(**data_dic)
                        result = ak.get_text(response.text, value[7])

                        try:
                            if result == value[8]:
                                # sheet.cell(r, 10).value = "通过"
                                print("通过")
                            else:

                                sheet.cell(r, 10).value = "不通过"
                                print("不通过：result是{}，期望是：{}".format(result, value[8]))
                        except Exception as e:
                            print('error:', e)
                else:
                    # print("不存在请求头")
                    if value[5]:
                        data_dic = {
                            "url": value[1] + value[2],
                            value[6]: eval(value[5])
                        }
                        response = getattr(ak, value[3])(**data_dic)
                        result = ak.get_text(response.text, value[7])

                        try:
                            if result == value[8]:
                                # sheet.cell(r, 10).value = "通过"
                                print("通过")
                            else:

                                sheet.cell(r, 10).value = "不通过"
                                print("不通过：result是{}，期望是：{}".format(result, value[8]))
                        except Exception as e:
                            print('error:', e)
                    else:
                        response = getattr(ak, value[3])(**data_dic)
                        result = ak.get_text(response.text, value[7])

                        try:
                            if result == value[8]:
                                # sheet.cell(r, 10).value = "通过"
                                print("通过")
                            else:

                                sheet.cell(r, 10).value = "不通过"
                                print("不通过：result是{}，期望是：{}".format(result, value[8]))
                        except Exception as e:
                            print('error:', e)
            else:
                print("我不是case")



